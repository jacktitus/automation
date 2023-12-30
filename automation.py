'''
the program is to automate the excelsheet using python we need to reduce the price 10% .maybe it's a human error or machine error
so if we want to solve this we cant just waste our time my entering the correct price manually,in this excel we have only 4 rows so
it's easy to modify by using formula in excel (select the cell *0.9) u can get the value, but if have an 1000 of rows we can't enter
it manually, so we are using pyhon to automate itself within seconds
'''

'''
import openpyxl as xl                      # xl here is called as an alias ,short our code
wb = xl.load_workbook('transactions.xlsx')
sheet = a['Sheet1']                        #access the excelsheet which u want
cell =sheet['a1']                          # coordinate of the column are value a1 means 1st row and 1st column
#print(cell.value)                          # print(sheet.max_row)-> this command will telll us how many rows are there in our excel

#for loop to access all the rows
for row in range(2, sheet.max_row +1):      # the reason why we are using +1 is ,the range function will take only 1-3 it omits the last value.
    cell = sheet.cell(row,3)                 # accessing the price column (we have to reduce it to 10%)
    #print(cell.value)
    corrected_price = cell.value * 0.9        #we are *0.9 for correct price
    corrected_price_cell = sheet.cell(row,4)     #making new column to store it
    corrected_price_cell.value = corrected_price  #adding values to the new column

wb.save('transations2.xlsx')                 # saving as a seperate file beacuse if we overwrite it may have some bug
'''

'''
# visual
import openpyxl as xl
from openpyxl.chart import BarChart,Reference,PieChart3D
wb = xl.load_workbook('transactions.xlsx')
sheet = wb['Sheet1']

for row in range(2, sheet.max_row +1):
    cell = sheet.cell(row,3)
    corrected_price = cell.value * 0.9
    corrected_price_cell = sheet.cell(row,4)
    corrected_price_cell.value = corrected_price
# visualizing
ref = Reference(sheet,
                min_col=4, max_col=4, min_row=2, max_row=sheet.max_row)
chart = PieChart3D()
chart.add_data(ref)
sheet.add_chart(chart,'e2')

wb.save('transations2.xlsx')
'''

# using function for resue the code just by entering the file name
import openpyxl as xl
from openpyxl.chart import BarChart, Reference, PieChart3D


def workbook(filename):
    wb = xl.load_workbook('filename')
    sheet = wb['Sheet1']

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row, 3)
        corrected_price = cell.value * 0.9
        corrected_price_cell = sheet.cell(row, 4)
        corrected_price_cell.value = corrected_price
    # visualizing
    ref = Reference(sheet,
                    min_col=4, max_col=4, min_row=2, max_row=sheet.max_row)
    chart = PieChart3D()
    chart.add_data(ref)
    sheet.add_chart(chart, 'e2')

    wb.save('filename')


workbook('filename')  # call the function and enter your file name
